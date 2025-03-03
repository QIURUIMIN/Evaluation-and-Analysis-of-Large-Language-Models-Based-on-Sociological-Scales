#!user/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
from flask import Flask, render_template, request, Response
import sqlite3
from create_tables import create_tables
from database import get_db_conn
import os


from ws4py.server.geventserver import WebSocketWSGIApplication
from ws4py.websocket import EchoWebSocket
from gevent import pywsgi
import json
import threading
import requests


app = Flask(__name__)

# 创建数据库表结构
create_tables()

from flask import Flask, request, Response, render_template
Uploaded_scales = []


def get_uploaded_scales():
    conn = get_db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT name, description, file_path FROM scales")
    uploaded_scales = cursor.fetchall()
    conn.close()
    return uploaded_scales



def is_file_duplicate(filename):
    conn = get_db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM scales WHERE file_path = ?", ("static/uploads/" + filename,))
    count = cursor.fetchone()[0]
    conn.close()
    return count > 0



@app.route("/upload_scale", methods=["POST"])
def upload_scale():
    if request.method == "POST":
        scale_name = request.form["scale-name"]
        scale_description = request.form["scale-description"]
        scale_file = request.files["scale-file"]

        #检查文件是否存在
        if is_file_duplicate(scale_file.filename):
            return "文件已存在，请更改文件名后重新上传！"

        # 将上传的文件保存到服务器上，具体路径可根据需求修改
        scale_file.save("static/uploads/" + scale_file.filename)
        print("上传成功！")

        # 将上传的量表信息存储到数据库中
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO scales (name, description, file_path) VALUES (?, ?, ?)",
                       (scale_name, scale_description, "static/uploads/" + scale_file.filename))
        conn.commit()
        conn.close()
        print("上传的量表信息已成功存储到数据库中！")

        # 在上传成功后，查询数据库获取已上传的量表信息
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT name, description, file_path FROM scales")
        uploaded_scales = cursor.fetchall()
        conn.close()
        print("已上传的量表信息：", uploaded_scales)

        # 打印当前工作目录
        import os
        print("当前工作目录：", os.getcwd())



        print(uploaded_scales)
        # 将新上传的量表信息添加到已有的量表列表中
        #new_scale = {"name": scale_name, "description": scale_description, "file_path": "static/uploads/" + scale_file.filename}
        #Uploaded_scales.append(new_scale)


        # 渲染模板并将更新后的量表列表传递给模板
        return render_template("Scales3.html", uploaded_scales=uploaded_scales)






@app.route("/")
def index():
    with open("static/Homepage2.html", encoding='utf-8') as f:
        content = f.read()
    return content

@app.route("/Scales2")
def Scales():
    uploaded_scales = get_uploaded_scales()  # 从数据库中获取已上传的量表
    print(uploaded_scales)
    return render_template("Scales3.html", uploaded_scales=uploaded_scales)

@app.route("/LLM")
def LLM():
    with open("static/LLM.html", encoding='utf-8') as f:
        content = f.read()
    return content

@app.route("/Interactive")
def Interactive():
    with open("static/Interactive2.html", encoding='utf-8') as f:
        content = f.read()
    return content

@app.route("/Visualization")
def Visualization():
    with open("static/Visualization.html", encoding='utf-8') as f:
        content = f.read()
    return content


from flask import Flask, request, jsonify
import API.Spark as spark  # 假设你的大模型处理函数在spark.py文件中
@app.route('/API', methods=['POST'])
def handle_query():
    try:
        data = request.get_json()
        print("Data:",data)
        question = data.get('question')
        print("Question:",question)
        model_id = data.get('modelId', 'default_model')  # 提供默认值
        print("Model ID:", model_id)

        answer = spark.api(question, model_id)  # 假设 api 函数接收 model_id
        print("Answer:", answer)
        return jsonify({'answer': answer})
    except Exception as e:
        app.logger.error(f"Error: {e}")  # 记录错误日志
        return jsonify({'error': str(e)}), 500
#运行应用程序

def api_test():
    answer = spark.api("hello")
    return answer



@app.route('/get_mbti_data/<type>')
def get_mbti_data(type):
    base_path = 'C:/Users/14542/Desktop/MProject/scales/'
    values_path = f'{base_path}{type}_Extracted_Values.xlsx'
    scores_path = f'{base_path}{type}_Extracted_Values_Personality_Scores.xlsx'
    data = {'values_data': [], 'scores_data': []}

    # 读取基本数据文件
    if os.path.exists(values_path):
        wb_values = openpyxl.load_workbook(values_path)
        sheet_values = wb_values.active
        for row in sheet_values.iter_rows(values_only=True):
            data['values_data'].append(row)

    # 读取个性得分数据文件
    if os.path.exists(scores_path):
        wb_scores = openpyxl.load_workbook(scores_path)
        sheet_scores = wb_scores.active
        for row in sheet_scores.iter_rows(values_only=True):
            data['scores_data'].append(row)

    return jsonify(data)


if __name__ == "__main__":
    app.run()

#查看当前flask的所有路由以及函数对应的关系（这其实就是路由）
#print(app.url_map)

